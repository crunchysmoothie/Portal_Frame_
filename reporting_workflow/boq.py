"""Structural-steel quantity take-off and tender BOQ workbook export."""

from __future__ import annotations

from collections import defaultdict
from copy import copy
import csv
from datetime import datetime
import math
from pathlib import Path
from typing import Any, Mapping, Sequence

from databases import member_database as mdb


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_BOQ_TEMPLATE = (
    PROJECT_ROOT
    / "references"
    / "BoQ"
    / "1. For Project Estimate"
    / "3. Structural Steelwork"
    / "Nafasi Water -Arnot Waste Treatment Plant - Structural Steel BoQ_T1.xlsx"
)
STEEL_DENSITY_KG_M3 = 7_850.0
ALLOWED_ADDITIONAL_UNITS = ("t", "kg", "m", "m2", "m3", "No", "Sum")


def _finite_number(
    value: Any,
    label: str,
    *,
    minimum: float = 0.0,
) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} must be a number.") from exc
    if not math.isfinite(number) or number < minimum:
        raise ValueError(f"{label} must be at least {minimum:g}.")
    return number


def normalise_additional_items(
    raw_items: Sequence[Mapping[str, Any]] | None,
) -> list[dict[str, Any]]:
    """Validate user-entered BOQ items without interpreting their scope."""

    items: list[dict[str, Any]] = []
    for index, raw in enumerate(raw_items or (), 1):
        description = str(raw.get("description", "")).strip()
        quantity_text = str(raw.get("quantity", "")).strip()
        rate_text = str(raw.get("rate", "")).strip()
        if not description and not quantity_text and not rate_text:
            continue
        if not description:
            raise ValueError(f"Additional item {index} requires a description.")
        unit = str(raw.get("unit", "")).strip()
        if unit not in ALLOWED_ADDITIONAL_UNITS:
            raise ValueError(
                f"Additional item {index} unit must be one of: "
                f"{', '.join(ALLOWED_ADDITIONAL_UNITS)}."
            )
        items.append({
            "description": description,
            "unit": unit,
            "quantity": _finite_number(
                raw.get("quantity"), f"Additional item {index} quantity"
            ),
            "rate": (
                _finite_number(raw.get("rate"), f"Additional item {index} rate")
                if rate_text
                else 0.0
            ),
            "source": "User-entered additional BOQ item",
        })
    return items


def _auxiliary_mass_lookup() -> dict[tuple[str, str], float]:
    lookup: dict[tuple[str, str], float] = {}
    database_path = PROJECT_ROOT / "databases" / "bracing_member_database.csv"
    with database_path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            family = str(row.get("section_type", "")).strip()
            designation = str(row.get("Designation", "")).strip()
            mass = str(row.get("m", "") or "").strip()
            if family and designation and mass:
                lookup[(family, designation)] = float(mass)
    return lookup


def _section_mass_kg_m(
    family: str,
    designation: str,
    rolled: Mapping[str, Mapping[str, Mapping[str, Any]]],
    auxiliary: Mapping[tuple[str, str], float],
) -> float:
    if family in rolled and designation in rolled[family]:
        return float(rolled[family][designation]["m"])
    direct = auxiliary.get((family, designation))
    if direct is not None:
        return float(direct)
    matches = [
        mass for (candidate_family, candidate), mass in auxiliary.items()
        if candidate == designation
    ]
    if len(matches) == 1:
        return float(matches[0])
    raise ValueError(
        f"Mass per metre was not found for {family} {designation}."
    )


def _mass_item(
    description: str,
    mass_kg: float,
    *,
    length_m: float | None = None,
    mass_per_m_kg: float | None = None,
    source: str,
) -> dict[str, Any]:
    item = {
        "description": description,
        "unit": "t",
        "quantity": max(float(mass_kg), 0.0) / 1000.0,
        "rate": 0.0,
        "mass_kg": max(float(mass_kg), 0.0),
        "source": source,
    }
    if length_m is not None:
        item["length_m"] = float(length_m)
    if mass_per_m_kg is not None:
        item["mass_per_m_kg"] = float(mass_per_m_kg)
    return item


def _girt_total_length_m(project: Mapping[str, Any]) -> float:
    if str(project.get("building_type", "")).lower() == "canopy":
        return 0.0
    length = float(project.get("building_length_mm", 0.0)) / 1000.0
    span = float(project.get("gable_width_mm", 0.0)) / 1000.0
    eaves = float(project.get("eaves_height_mm", 0.0)) / 1000.0
    apex = float(project.get("apex_height_mm", 0.0)) / 1000.0
    maximum = float(project.get("girt_max_spacing_mm", 0.0)) / 1000.0
    if min(length, span, eaves, maximum) <= 0.0:
        return 0.0

    side_spaces = max(1, math.ceil(eaves / maximum))
    side_length = 2.0 * length * side_spaces

    roof_type = str(project.get("roof_type", "Duo Pitched"))
    full_height = max(apex, eaves)
    gable_spaces = max(1, math.ceil(full_height / maximum))
    actual = full_height / gable_spaces
    gable_length_one_end = 0.0
    for index in range(1, gable_spaces + 1):
        height = index * actual
        if roof_type == "Duo Pitched" and height > eaves and apex > eaves:
            width = span * max(0.0, (apex - height) / (apex - eaves))
        elif roof_type == "Mono Pitched" and height > eaves and apex > eaves:
            width = span * max(0.0, (apex - height) / (apex - eaves))
        else:
            width = span
        gable_length_one_end += width
    return side_length + 2.0 * gable_length_one_end


def _plate_and_bolt_items(
    connection_design: Mapping[str, Any],
    frame_count: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    plate_groups: dict[tuple[str, float, float, float], int] = defaultdict(int)
    bolt_groups: dict[tuple[str, float, str], int] = defaultdict(int)

    for support in connection_design.get("base_plates", {}).get("supports", []):
        plate = support.get("plate", {})
        length = float(plate.get("length_mm", 0.0) or 0.0)
        width = float(plate.get("width_mm", 0.0) or 0.0)
        thickness = float(plate.get("provided_thickness_mm", 0.0) or 0.0)
        if min(length, width, thickness) > 0.0:
            plate_groups[("Base plate", length, width, thickness)] += frame_count
        bolts = support.get("holding_down_bolts", {})
        layout = bolts.get("layout", {})
        diameter = float(layout.get("diameter_mm", 0.0) or 0.0)
        count = int(layout.get("bolt_count", 0) or 0)
        if diameter > 0.0 and count > 0:
            grade = str(bolts.get("steel_grade", "8.8"))
            bolt_groups[(grade, diameter, "holding-down bolts")] += count * frame_count

    for location in connection_design.get("haunch_connections", {}).get("locations", []):
        connection = location.get("connection", {})
        plate = connection.get("plate", {})
        height = float(plate.get("height_mm", 0.0) or 0.0)
        width = float(plate.get("width_mm", 0.0) or 0.0)
        thickness = float(plate.get("provided_thickness_mm", 0.0) or 0.0)
        label = (
            "Apex end plate"
            if str(location.get("connection_type")) == "apex_splice"
            else "Eaves end plate"
        )
        if min(height, width, thickness) > 0.0:
            plate_groups[(label, height, width, thickness)] += frame_count
        bolts = connection.get("bolts", {})
        diameter = float(bolts.get("diameter_mm", 0.0) or 0.0)
        count = int(bolts.get("bolt_count", 0) or 0)
        if diameter > 0.0 and count > 0:
            bolt_groups[("8.8", diameter, f"{label.lower()} bolts")] += count * frame_count

    plate_items: list[dict[str, Any]] = []
    for (label, length, width, thickness), count in sorted(plate_groups.items()):
        mass = (
            length / 1000.0
            * width / 1000.0
            * thickness / 1000.0
            * STEEL_DENSITY_KG_M3
            * count
        )
        plate_items.append(_mass_item(
            f"{label}s {length:g} x {width:g} x {thickness:g} mm ({count} No)",
            mass,
            source="Calculated connection plate geometry and whole-building count",
        ))

    bolt_items = [
        {
            "description": (
                f"Grade {grade} M{diameter:g} {purpose}, nuts and washers"
            ),
            "unit": "No",
            "quantity": float(count),
            "rate": 0.0,
            "source": "Calculated connection layout and whole-building count",
        }
        for (grade, diameter, purpose), count in sorted(bolt_groups.items())
    ]
    return plate_items, bolt_items


def _cladding_items(project: Mapping[str, Any]) -> list[dict[str, Any]]:
    if str(project.get("building_type", "")).lower() == "canopy":
        return []
    length = float(project.get("building_length_mm", 0.0)) / 1000.0
    span = float(project.get("gable_width_mm", 0.0)) / 1000.0
    eaves = float(project.get("eaves_height_mm", 0.0)) / 1000.0
    apex = float(project.get("apex_height_mm", 0.0)) / 1000.0
    rise = max(apex - eaves, 0.0)
    roof_type = str(project.get("roof_type", "Duo Pitched"))
    run = span / 2.0 if roof_type == "Duo Pitched" else span
    slope = math.hypot(run, rise)
    roof_area = (2.0 * slope if roof_type == "Duo Pitched" else slope) * length
    if roof_type == "Duo Pitched":
        gable_area_each = span * eaves + span * rise / 2.0
        total_barge = 4.0 * slope
    else:
        gable_area_each = span * (eaves + apex) / 2.0
        total_barge = 2.0 * slope
    wall_area = 2.0 * length * eaves + 2.0 * gable_area_each
    openings = project.get("wall_openings_m2", {})
    if isinstance(openings, Mapping):
        wall_area -= sum(float(value or 0.0) for value in openings.values())
    wall_area = max(wall_area, 0.0)

    items = [
        {
            "description": (
                "0,8mm IBR 686 prepainted inverted boxed rib sheeting with one "
                "stiffener rib in each pan, and with \"Colorplus Ultimate Grey\" "
                "finish of AZ200 on one side in single lengths, fixed to steel "
                "purlins with stainless steel top speed screws with galvanised "
                "EPDM bonded washers in strict accordance with the manufacturer's specifications."
            ),
            "unit": "m2",
            "quantity": roof_area,
            "rate": 0.0,
            "source": "Calculated roof slope area",
        },
        {
            "description": (
                "0,6mm IBR 686 prepainted inverted boxed rib sheeting with one "
                "stiffener rib in each pan, and with \"Colorplus Ultimate Grey\" "
                "finish of AZ200 on one side in single lengths, fixed to steel "
                "girts with stainless steel top speed screws with galvanised "
                "EPDM bonded washers in strict accordance with the manufacturer's specifications."
            ),
            "unit": "m2",
            "quantity": wall_area,
            "rate": 0.0,
            "source": "Calculated wall and gable area less entered openings",
        },
    ]
    items.append({
        "description": "Extra over roof covering for 82 degrees bullnose to 450mm radius at eaves edge",
        "unit": "m",
        "quantity": 2.0 * length,
        "rate": 0.0,
        "source": "Reference BOQ assumption: bullnose at both eaves edges",
    })
    if roof_type == "Duo Pitched":
        items.append({
            "description": "0,8mm IBR 686 Ridge cap flashing (600mm girth) with \"Colorplus Ultimate Grey\" finish of AZ200 complete with serrated closers and poly ridge closers",
            "unit": "m",
            "quantity": length,
            "rate": 0.0,
            "source": "Calculated building ridge length",
        })
    items.extend([
        {
            "description": "0,8mm IBR 686 Gable end flashing with \"Colorplus Ultimate Grey\" finish of AZ200 complete",
            "unit": "m",
            "quantity": 2.0 * span,
            "rate": 0.0,
            "source": "Reference BOQ assumption: two gable ends",
        },
        {
            "description": "0,6mm IBR 686 External corner flashing with \"Colorplus Ultimate Grey\" finish of AZ200 complete",
            "unit": "m",
            "quantity": 4.0 * eaves,
            "rate": 0.0,
            "source": "Reference BOQ assumption: four external corners",
        },
        {
            "description": "0,6mm IBR 686 Drip detail flashing with \"Colorplus Ultimate Grey\" finish of AZ200 complete",
            "unit": "m",
            "quantity": 2.0 * length,
            "rate": 0.0,
            "source": "Reference BOQ assumption: both eaves edges",
        },
        {
            "description": "0,6mm IBR 686 Gable barge flashing with \"Colorplus Ultimate Grey\" finish of AZ200 complete",
            "unit": "m",
            "quantity": total_barge,
            "rate": 0.0,
            "source": "Calculated roof-edge length at both gable ends",
        },
    ])
    return items


def build_structural_boq_takeoff(
    snapshot: Mapping[str, Any],
    connection_design: Mapping[str, Any],
    additional_items: Sequence[Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    """Build an auditable whole-building structural-steel take-off."""

    results = snapshot.get("results", {})
    project = dict(results.get("project", {}))
    summary = dict(results.get("frame_summary", {}))
    bracing = dict(results.get("bracing_design", {}))
    mass = dict(summary.get("steel_mass_breakdown", {}))
    frame_count = int(mass.get("portal_frames", {}).get("quantity", 0) or 0)
    if frame_count <= 0:
        length = float(project.get("building_length_mm", 0.0))
        spacing = float(project.get("rafter_spacing_mm", 0.0))
        if min(length, spacing) <= 0.0:
            raise ValueError("The completed analysis does not contain a frame count.")
        frame_count = math.ceil(length / spacing - 1e-9) + 1

    rolled = mdb.load_member_database()
    auxiliary = _auxiliary_mass_lookup()
    steel_items: list[dict[str, Any]] = []

    column_section = str(project.get("column_section", "")).strip()
    rafter_section = str(project.get("rafter_section", "")).strip()
    input_frame = snapshot.get("input_data", {}).get("frame_data", [{}])[0]
    column_family = str(input_frame.get("column_section_type", "H-Sections"))
    rafter_family = str(input_frame.get("rafter_section_type", "I-Sections"))
    column_m = _section_mass_kg_m(column_family, column_section, rolled, auxiliary)
    rafter_m = _section_mass_kg_m(rafter_family, rafter_section, rolled, auxiliary)
    column_length = float(summary.get("column_length_m", 0.0)) * frame_count
    rafter_length = float(summary.get("rafter_length_m", 0.0)) * frame_count
    steel_items.extend([
        _mass_item(
            f"Columns - {column_section}",
            column_length * column_m,
            length_m=column_length,
            mass_per_m_kg=column_m,
            source="Analysed column length x frame count x section mass",
        ),
        _mass_item(
            f"Rafters - {rafter_section}",
            rafter_length * rafter_m,
            length_m=rafter_length,
            mass_per_m_kg=rafter_m,
            source="Analysed rafter length x frame count x section mass",
        ),
    ])
    haunch_mass = float(summary.get("haunch_mass_per_frame_kg", 0.0)) * frame_count
    if haunch_mass > 0.0:
        steel_items.append(_mass_item(
            f"Haunches cut from {rafter_section}",
            haunch_mass,
            source="Calculated tapered-haunch extra steel x frame count",
        ))

    gable_groups: dict[tuple[str, str], float] = defaultdict(float)
    for column in bracing.get("gable_columns", []):
        key = (str(column.get("section_type", "")), str(column.get("section", "")))
        gable_groups[key] += float(column.get("height_mm", 0.0)) / 1000.0 * 2.0
    for (family, designation), total_length in sorted(gable_groups.items()):
        kg_m = _section_mass_kg_m(family, designation, rolled, auxiliary)
        steel_items.append(_mass_item(
            f"Gable columns - {designation}",
            total_length * kg_m,
            length_m=total_length,
            mass_per_m_kg=kg_m,
            source="Calculated gable-column heights at both ends",
        ))

    braced_bay_count = int(mass.get("bracing", {}).get("braced_bay_count", 0) or 0)
    brace_by_type = {
        str(item.get("member_type", "")): item
        for item in bracing.get("bracing_members", [])
    }
    roof_brace = brace_by_type.get("Roof X-brace")
    roof_points = bracing.get("roof_layout", {}).get("roof_points", [])
    panels = bracing.get("roof_layout", {}).get("brace_panels", [])
    bay_m = float(bracing.get("roof_layout", {}).get("bay_length_mm", 0.0)) / 1000.0
    if roof_brace and len(roof_points) >= 2 and braced_bay_count:
        if not panels:
            panels = [
                {"start_index": index, "end_index": index + 1}
                for index in range(len(roof_points) - 1)
            ]
        length_one_bay = 2.0 * sum(
            math.hypot(
                bay_m,
                math.hypot(
                    float(roof_points[item["end_index"]]["x_mm"])
                    - float(roof_points[item["start_index"]]["x_mm"]),
                    float(roof_points[item["end_index"]]["y_mm"])
                    - float(roof_points[item["start_index"]]["y_mm"]),
                ) / 1000.0,
            )
            for item in panels
        )
        total_length = length_one_bay * braced_bay_count
        family = str(roof_brace.get("section_family", ""))
        designation = str(roof_brace.get("section", ""))
        kg_m = _section_mass_kg_m(family, designation, rolled, auxiliary)
        steel_items.append(_mass_item(
            f"Roof X-bracing - {designation}",
            total_length * kg_m,
            length_m=total_length,
            mass_per_m_kg=kg_m,
            source="Calculated X-brace diagonals in the selected braced bays",
        ))

    side_brace = brace_by_type.get("Longitudinal side-wall brace")
    if side_brace and braced_bay_count:
        members_per_wall = int(
            bracing.get("column_bracing_layout", {}).get("members_per_wall", 1)
        )
        total_length = (
            float(side_brace.get("length_mm", 0.0)) / 1000.0
            * members_per_wall
            * 2.0
            * braced_bay_count
        )
        family = str(side_brace.get("section_family", ""))
        designation = str(side_brace.get("section", ""))
        kg_m = _section_mass_kg_m(family, designation, rolled, auxiliary)
        steel_items.append(_mass_item(
            f"Longitudinal wall bracing - {designation}",
            total_length * kg_m,
            length_m=total_length,
            mass_per_m_kg=kg_m,
            source="Calculated wall-brace members on both long walls",
        ))

    purlin = mass.get("purlins", {})
    purlin_section = str(purlin.get("section", project.get("purlin_section", "")))
    purlin_length = float(purlin.get("total_length_m", 0.0) or 0.0)
    if purlin_section and purlin_length > 0.0:
        kg_m = _section_mass_kg_m("Lipped Channels", purlin_section, rolled, auxiliary)
        steel_items.append(_mass_item(
            f"Purlins - {purlin_section}",
            purlin_length * kg_m,
            length_m=purlin_length,
            mass_per_m_kg=kg_m,
            source="Calculated purlin lines x building length",
        ))

    girt_section = str(project.get("girt_section", "")).strip()
    girt_length = _girt_total_length_m(project)
    if girt_section and girt_length > 0.0:
        kg_m = _section_mass_kg_m("Lipped Channels", girt_section, rolled, auxiliary)
        steel_items.append(_mass_item(
            f"Girts - {girt_section}",
            girt_length * kg_m,
            length_m=girt_length,
            mass_per_m_kg=kg_m,
            source="Calculated girt rows around side and gable walls",
        ))

    plate_items, bolt_items = _plate_and_bolt_items(connection_design, frame_count)
    steel_items.extend(plate_items)
    fabricated_mass_t = sum(
        float(item["quantity"]) for item in steel_items if item["unit"] == "t"
    )
    # The supplied tender BOQs use these commercial allowances consistently:
    # erection bolts are 4.5% of delivered steel mass, and erection includes
    # both the steel and that bolt allowance.  Keep the detailed connection
    # bolt count above for audit, while pricing the reference-BoQ tonne item.
    erection_bolt_mass_t = fabricated_mass_t * 0.045
    reference_items = [
        {
            "description": "Preparation of shop detail drawings",
            "unit": "t",
            "quantity": fabricated_mass_t,
            "rate": 0.0,
            "source": "Reference BOQ assumption: quantity equals fabricated steel tonnage",
        },
        {
            "description": "Normal delivery",
            "unit": "t",
            "quantity": fabricated_mass_t,
            "rate": 0.0,
            "source": "Reference BOQ assumption: quantity equals fabricated steel tonnage",
        },
        {
            "description": "Structural steelwork and erection bolts",
            "unit": "t",
            "quantity": fabricated_mass_t + erection_bolt_mass_t,
            "rate": 0.0,
            "source": "Reference BOQ assumption: steel tonnage plus 4.5% erection-bolt allowance",
        },
        {
            "description": "Grade 8.8 precision galvanised bolts, nuts and washers",
            "unit": "t",
            "quantity": erection_bolt_mass_t,
            "rate": 0.0,
            "source": "Reference BOQ assumption: erection bolts equal 4.5% of delivered steel tonnage",
        },
        {
            "description": "Corrosion protection for all structural steel members must be Hot-Dipped Galvanising applied according to SANS 121 (ISO 1461)",
            "unit": "t",
            "quantity": fabricated_mass_t,
            "rate": 0.0,
            "source": "Reference BOQ assumption: galvanising quantity equals structural steel tonnage",
        },
    ]
    return {
        "schema_version": 1,
        "generated": datetime.now().astimezone().isoformat(timespec="seconds"),
        "analysis_id": str(snapshot.get("analysis", {}).get("analysis_id", "")),
        "project": {
            "name": str(project.get("project_name", "Untitled project")),
            "number": str(project.get("project_number", "")),
            "designer": str(project.get("designer", "")),
        },
        "frame_count": frame_count,
        "fabricated_steel_mass_t": fabricated_mass_t,
        "steel_items": steel_items,
        "bolt_items": bolt_items,
        "reference_items": reference_items,
        "cladding_items": _cladding_items(project),
        "additional_items": normalise_additional_items(additional_items),
        "assumptions": [
            "Calculated quantities use the completed PortalFrame analysis snapshot.",
            "Rates are user-editable and default to zero.",
            "Cladding follows the supplied BOQ assumptions: 0.8mm roof IBR, 0.6mm side/gable IBR, Colorplus Ultimate Grey AZ200, 82-degree bullnose to 450mm radius, 600mm-girth ridge cap, gable/end, corner, drip and barge flashings.",
            "Girt quantities assume continuous rows around both side and gable walls; opening trimmers are excluded.",
            "Connection plate and bolt quantities repeat the calculated typical transverse-frame connections at every frame line.",
            "Preparation drawings, normal delivery, erection and galvanising follow the supplied BOQ tonnage assumptions; erection bolts are 4.5% of delivered steel mass.",
            "All quantities require engineer and quantity-surveyor review before tender issue.",
        ],
    }


def build_truss_structural_boq_takeoff(
    snapshot: Mapping[str, Any],
    connection_design: Mapping[str, Any],
    additional_items: Sequence[Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    """Build the structural take-off for the selected truss arrangement."""

    results = snapshot.get("results", {})
    project = dict(results.get("project", {}))
    truss_result = results.get("truss_design", {})
    ranked = list(truss_result.get("ranked_solutions", []))
    if not ranked:
        raise ValueError("The completed truss analysis has no ranked solution.")
    best = ranked[0]
    frame_count = int(best.get("truss_count", 0) or 0)
    if frame_count <= 0:
        raise ValueError("The completed truss analysis has no truss-line count.")

    steel_items: list[dict[str, Any]] = []
    role_labels = {
        "top_chord": "Truss top chords",
        "bottom_chord": "Truss bottom chords",
        "diagonal": "Truss diagonal webs",
        "vertical": "Truss vertical webs",
        "support_vertical": "Truss bearing verticals",
    }
    member_groups: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {"length_m": 0.0, "mass_kg": 0.0}
    )
    for member in best.get("member_schedule", []):
        role = str(member.get("role", ""))
        section = member.get("section", {})
        designation = str(section.get("designation", "")).strip()
        if not designation:
            continue
        length_m = float(member.get("length_mm", 0.0)) / 1000.0 * frame_count
        mass_per_m = float(section.get("mass_kg_m", 0.0) or 0.0)
        group = member_groups[(role, designation)]
        group["length_m"] += length_m
        group["mass_kg"] += length_m * mass_per_m
        group["mass_per_m_kg"] = mass_per_m
    for (role, designation), values in sorted(member_groups.items()):
        steel_items.append(_mass_item(
            f"{role_labels.get(role, role.replace('_', ' ').title())} - {designation}",
            values["mass_kg"],
            length_m=values["length_m"],
            mass_per_m_kg=values["mass_per_m_kg"],
            source="Selected truss member lengths x truss-line count x exact section mass",
        ))

    eave = best.get("eave_column_design", {})
    if eave.get("status") == "PASS":
        steel_items.append(_mass_item(
            f"Main eave columns - {eave.get('section', '')}",
            float(eave.get("total_mass_kg", 0.0)),
            length_m=(
                float(eave.get("height_mm", 0.0)) / 1000.0
                * int(eave.get("column_count", 0) or 0)
            ),
            mass_per_m_kg=float(eave.get("mass_kg_m", 0.0)),
            source="Calculated eave-column height x whole-building column count",
        ))
    centre = best.get("centre_column_design", {})
    if centre.get("status") == "PASS":
        steel_items.append(_mass_item(
            f"Centre columns - {centre.get('section', '')}",
            float(centre.get("total_mass_kg", 0.0)),
            length_m=(
                float(centre.get("height_mm", 0.0)) / 1000.0
                * int(centre.get("column_count", 0) or 0)
            ),
            mass_per_m_kg=float(centre.get("mass_kg_m", 0.0)),
            source="Calculated centre-column height x whole-building column count",
        ))

    girder = best.get("girder_design", {})
    if girder.get("status") == "PASS":
        repeated = int(girder.get("repeated_span_count", 0) or 0)
        girder_groups: dict[tuple[str, str], dict[str, float]] = defaultdict(
            lambda: {"length_m": 0.0, "mass_kg": 0.0}
        )
        for member in girder.get("member_schedule", []):
            role = str(member.get("role", ""))
            section = member.get("section", {})
            designation = str(section.get("designation", "")).strip()
            length_m = float(member.get("length_mm", 0.0)) / 1000.0 * repeated
            mass_per_m = float(section.get("mass_kg_m", 0.0) or 0.0)
            group = girder_groups[(role, designation)]
            group["length_m"] += length_m
            group["mass_kg"] += length_m * mass_per_m
            group["mass_per_m_kg"] = mass_per_m
        for (role, designation), values in sorted(girder_groups.items()):
            steel_items.append(_mass_item(
                f"Longitudinal lattice-girder {role.replace('_', ' ')} - {designation}",
                values["mass_kg"],
                length_m=values["length_m"],
                mass_per_m_kg=values["mass_per_m_kg"],
                source="Selected girder member lengths x repeated girder spans",
            ))

    purlin = best.get("purlins", {})
    if float(purlin.get("mass_kg", 0.0) or 0.0) > 0.0:
        steel_items.append(_mass_item(
            f"Purlins - {purlin.get('section', '')}",
            float(purlin["mass_kg"]),
            length_m=float(purlin.get("total_length_m", 0.0)),
            mass_per_m_kg=float(purlin.get("mass_per_m_kg", 0.0)),
            source="Calculated purlin lines x building length",
        ))

    rolled = mdb.load_member_database()
    auxiliary = _auxiliary_mass_lookup()
    gable_groups: dict[tuple[str, str], float] = defaultdict(float)
    for column in results.get("bracing_design", {}).get("gable_columns", []):
        key = (str(column.get("section_type", "")), str(column.get("section", "")))
        gable_groups[key] += float(column.get("height_mm", 0.0)) / 1000.0 * 2.0
    for (family, designation), length_m in sorted(gable_groups.items()):
        mass_per_m = _section_mass_kg_m(family, designation, rolled, auxiliary)
        steel_items.append(_mass_item(
            f"Gable columns - {designation}",
            length_m * mass_per_m,
            length_m=length_m,
            mass_per_m_kg=mass_per_m,
            source="Calculated gable-column heights at both ends",
        ))

    girt_section = str(project.get("girt_section", "")).strip()
    girt_length = _girt_total_length_m(project)
    if girt_section and girt_length > 0.0:
        mass_per_m = _section_mass_kg_m(
            "Lipped Channels", girt_section, rolled, auxiliary
        )
        steel_items.append(_mass_item(
            f"Girts - {girt_section}",
            girt_length * mass_per_m,
            length_m=girt_length,
            mass_per_m_kg=mass_per_m,
            source="Calculated girt rows around side and gable walls",
        ))

    plate_items, bolt_items = _plate_and_bolt_items(connection_design, frame_count)
    steel_items.extend(plate_items)
    fabricated_mass_t = sum(
        float(item["quantity"]) for item in steel_items if item["unit"] == "t"
    )
    erection_bolt_mass_t = fabricated_mass_t * 0.045
    reference_items = [
        {"description": "Preparation of shop detail drawings", "unit": "t", "quantity": fabricated_mass_t, "rate": 0.0, "source": "Reference BOQ assumption"},
        {"description": "Normal delivery", "unit": "t", "quantity": fabricated_mass_t, "rate": 0.0, "source": "Reference BOQ assumption"},
        {"description": "Structural steelwork and erection bolts", "unit": "t", "quantity": fabricated_mass_t + erection_bolt_mass_t, "rate": 0.0, "source": "Steel tonnage plus 4.5% erection-bolt allowance"},
        {"description": "Grade 8.8 precision galvanised bolts, nuts and washers", "unit": "t", "quantity": erection_bolt_mass_t, "rate": 0.0, "source": "4.5% erection-bolt allowance"},
        {"description": "Corrosion protection for all structural steel members must be Hot-Dipped Galvanising applied according to SANS 121 (ISO 1461)", "unit": "t", "quantity": fabricated_mass_t, "rate": 0.0, "source": "Reference BOQ assumption"},
    ]
    return {
        "schema_version": 1,
        "generated": datetime.now().astimezone().isoformat(timespec="seconds"),
        "analysis_id": str(snapshot.get("analysis", {}).get("analysis_id", "")),
        "project": {
            "name": str(project.get("project_name", "Untitled project")),
            "number": str(project.get("project_number", "")),
            "designer": str(project.get("designer", "")),
        },
        "frame_count": frame_count,
        "fabricated_steel_mass_t": fabricated_mass_t,
        "steel_items": steel_items,
        "bolt_items": bolt_items,
        "reference_items": reference_items,
        "cladding_items": _cladding_items(project),
        "additional_items": normalise_additional_items(additional_items),
        "assumptions": [
            "Calculated quantities use the selected rank-1 truss arrangement and exact section designations.",
            "No portal haunch steel or haunch connections are included; only calculated column base plates and holding-down bolts are measured.",
            "Truss gussets, chord/web joint bolts and welds, splices, restraint ties and global roof/wall bracing quantities remain input-required until those connections and members are designed.",
            "Rates are user-editable and default to zero.",
            "All quantities require engineer and quantity-surveyor review before tender issue.",
        ],
    }


def _capture_row_style(sheet: Any, row: int) -> list[Any]:
    return [copy(sheet.cell(row, column)._style) for column in range(1, 8)]


def _apply_row_style(sheet: Any, row: int, styles: Sequence[Any]) -> None:
    for column, style in enumerate(styles, 1):
        sheet.cell(row, column)._style = copy(style)


def write_structural_boq_xlsx(
    takeoff: Mapping[str, Any],
    output_path: str | Path,
    *,
    template_path: str | Path = DEFAULT_BOQ_TEMPLATE,
) -> Path:
    """Write the calculated take-off in the supplied tender BOQ format."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment guard
        raise RuntimeError(
            "Structural BOQ export requires openpyxl from requirements.txt."
        ) from exc

    template_path = Path(template_path)
    if not template_path.is_file():
        raise FileNotFoundError(f"Structural BOQ template not found: {template_path}")
    workbook = load_workbook(template_path)
    template = workbook["Section 2 - RO Building"]
    row_styles = {
        "plain": _capture_row_style(template, 13),
        "heading": _capture_row_style(template, 15),
        "subheading": _capture_row_style(template, 18),
        "section": _capture_row_style(template, 9),
        "header": _capture_row_style(template, 5),
        "total": _capture_row_style(template, 53),
        "long": _capture_row_style(template, 65),
    }
    row_heights = {
        name: template.row_dimensions[row].height
        for name, row in {
            "plain": 13, "heading": 15, "subheading": 18,
            "section": 9, "header": 5, "total": 53, "long": 65,
        }.items()
    }
    for sheet in list(workbook.worksheets):
        if sheet is not template:
            workbook.remove(sheet)
    template.title = "Structural Steel BoQ"
    for merged in list(template.merged_cells.ranges):
        template.unmerge_cells(str(merged))
    template.delete_rows(1, template.max_row)
    template.sheet_view.showGridLines = True

    row = 1
    project = takeoff.get("project", {})
    project_title = str(project.get("name") or "PORTAL FRAME PROJECT").upper()
    template.cell(row, 1, project_title)
    template.cell(row, 7, "Revision T1")
    _apply_row_style(template, row, row_styles["heading"])
    row += 1
    template.cell(row, 1, "STRUCTURAL STEELWORK SCHEDULE OF QUANTITIES")
    _apply_row_style(template, row, row_styles["heading"])
    row += 1
    section_title = f"SECTION 1: {project_title}"
    template.cell(row, 1, section_title)
    template.cell(row, 7, datetime.now().strftime("%d/%b/%y"))
    _apply_row_style(template, row, row_styles["heading"])
    row += 1
    _apply_row_style(template, row, row_styles["plain"])
    row += 1

    headers = ["PAYMENT\nREFERS TO", "ITEM\nNO", "DESCRIPTION", "UNIT", "QUANTITY", "RATE", "AMOUNT"]
    for column, value in enumerate(headers, 1):
        template.cell(row, column, value)
    _apply_row_style(template, row, row_styles["header"])
    template.row_dimensions[row].height = row_heights["header"]
    row += 1

    item_number = 0
    amount_rows: list[int] = []

    def styled_row(kind: str = "plain") -> int:
        nonlocal row
        current = row
        _apply_row_style(template, current, row_styles[kind])
        if row_heights.get(kind) is not None:
            template.row_dimensions[current].height = row_heights[kind]
        row += 1
        return current

    def blank() -> None:
        styled_row("plain")

    def heading(text: str, *, green: bool = False, payment_ref: str = "") -> None:
        current = styled_row("section" if green else "heading")
        template.cell(current, 1, payment_ref or None)
        template.cell(current, 3, text)

    def item(entry: Mapping[str, Any], payment_ref: str = "") -> None:
        nonlocal item_number
        item_number += 1
        kind = "long" if len(str(entry.get("description", ""))) > 95 else "plain"
        current = styled_row(kind)
        values = [
            payment_ref or None,
            item_number,
            str(entry.get("description", "")),
            str(entry.get("unit", "")),
            float(entry.get("quantity", 0.0)),
            float(entry.get("rate", 0.0)),
        ]
        for column, value in enumerate(values, 1):
            template.cell(current, column, value)
        template.cell(current, 7, f"=E{current}*F{current}")
        amount_rows.append(current)
        blank()

    heading("STRUCTURAL STEELWORK", green=True, payment_ref="SANS\n1200H")
    blank()
    heading("Supply and fabrication:")
    reference_items = list(takeoff.get("reference_items", []))
    if reference_items:
        item(reference_items[0], "8.3.1.1")
    heading("Supply and fabrication of steelwork:")
    current = styled_row("subheading")
    template.cell(current, 3, "Joint by welding in shop and bolting on site:")
    blank()
    current = styled_row("subheading")
    template.cell(current, 3, "Calculated steel sections and platework")
    blank()
    for entry in takeoff.get("steel_items", []):
        item(entry, "8.3.1.2")

    heading("Delivery to site", payment_ref="8.3.2")
    if len(reference_items) > 1:
        item(reference_items[1], "8.3.2.1")
    heading("Erection on site", payment_ref="8.3.3")
    if len(reference_items) > 2:
        item(reference_items[2])
    heading("Erection bolts", payment_ref="8.3.4")
    if len(reference_items) > 3:
        item(reference_items[3], "8.3.4.1")
    # The reference BOQ treats galvanising as a separate measured section.
    heading("CORROSION PROTECTION OF STRUCTURAL STEELWORK", green=True, payment_ref="1200HC")
    if len(reference_items) > 4:
        item(reference_items[4], "8.2.1")

    carried_row = styled_row("total")
    template.cell(carried_row, 1, "SECTION 1 - CARRIED FORWARD")
    template.cell(carried_row, 7, "=SUM(" + ",".join(f"G{value}" for value in amount_rows) + ")")
    blank()
    header_two = styled_row("header")
    for column, value in enumerate(headers, 1):
        template.cell(header_two, column, value)
    brought_row = styled_row("total")
    template.cell(brought_row, 1, "SECTION 1 - BROUGHT FORWARD")
    template.cell(brought_row, 7, f"=G{carried_row}")

    second_amount_rows: list[int] = []
    amount_rows = second_amount_rows
    if takeoff.get("cladding_items"):
        heading("CLADDING AND SHEETING", green=True, payment_ref="1200HB")
        blank()
        heading("Supply and install cladding, sheeting and ancillaries:")
        blank()
        for entry in takeoff.get("cladding_items", []):
            item(entry, "8.2.2")
    if takeoff.get("additional_items"):
        heading("ADDITIONAL ITEMS", green=True)
        blank()
        for entry in takeoff.get("additional_items", []):
            item(entry)

    total_row = styled_row("total")
    template.cell(total_row, 1, "TOTAL SECTION 1 CARRIED TO SUMMARY")
    second_sum = "+".join(f"G{value}" for value in second_amount_rows) or "0"
    template.cell(total_row, 7, f"=G{brought_row}+{second_sum}")

    # Preserve the reference workbook's summary-row geometry.  The source
    # BOQs merge the payment-reference through rate columns on carried,
    # brought-forward and total rows; leaving those cells unmerged creates
    # visible vertical lines and changes the centred/left-aligned appearance.
    for merge_row in (carried_row, brought_row, total_row):
        template.merge_cells(start_row=merge_row, start_column=1, end_row=merge_row, end_column=6)

    template.column_dimensions["A"].width = 15
    template.column_dimensions["B"].width = 9
    template.column_dimensions["C"].width = 60
    template.column_dimensions["D"].width = 9
    template.column_dimensions["E"].width = 12
    template.column_dimensions["F"].width = 13
    template.column_dimensions["G"].width = 15
    template.freeze_panes = "A6"
    template.print_title_rows = "1:5"
    template.print_area = f"A1:G{total_row}"
    template.sheet_properties.pageSetUpPr.fitToPage = True
    template.page_setup.fitToWidth = 1
    template.page_setup.fitToHeight = 0
    template.page_setup.orientation = "portrait"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
