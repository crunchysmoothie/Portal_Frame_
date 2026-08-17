"""First civil/concrete BOQ take-off using the supplied civil BOQ template."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
CIVIL_TEMPLATE = (
    PROJECT_ROOT
    / "references"
    / "BoQ"
    / "1. For Project Estimate"
    / "2. Civil and Concrete Works"
    / "Nafasi Water - Arnot Waste Tratment Plant - Piling, Building and Civil Works - Rev T1.xlsx"
)


def build_civil_boq_takeoff(
    snapshot: Mapping[str, Any],
    foundation_design: Mapping[str, Any],
    inputs: Mapping[str, Any],
    connection_design: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    project = dict(snapshot.get("results", {}).get("project", {}))
    frame_count = int(
        snapshot.get("results", {}).get("frame_summary", {})
        .get("steel_mass_breakdown", {})
        .get("portal_frames", {})
        .get("quantity", 0)
        or 0
    )
    length_m = float(project.get("building_length_mm", 0.0)) / 1000.0
    span_m = float(project.get("gable_width_mm", 0.0)) / 1000.0
    area_m2 = float(inputs.get("surface_bed_area_m2", 0.0) or 0.0) or length_m * span_m
    thickness_mm = float(inputs.get("surface_bed_thickness_mm", 150.0) or 150.0)
    spacing_m = float(inputs.get("joint_spacing_m", 6.0) or 6.0)
    excavation_depth_m = float(inputs.get("excavation_below_surface_bed_m", 0.0) or 0.0)
    footing_backfill_m3 = float(inputs.get("concrete_footing_backfill_m3", 0.0) or 0.0)

    foundation_inputs = dict(foundation_design.get("inputs", {}))
    footing_length = float(foundation_inputs.get("length_m", 0.0) or 0.0)
    footing_width = float(foundation_inputs.get("width_m", 0.0) or 0.0)
    footing_depth = float(foundation_inputs.get("base_depth_m", 0.0) or 0.0)
    footing_volume = float(foundation_design.get("derived", {}).get("footing_volume_m3", 0.0) or 0.0)
    support_rows = list(foundation_design.get("supports", []))
    support_count = int(sum(
        int(item.get("quantity", 1) or 1) for item in support_rows
    )) if support_rows else 2

    joint_length_m = 2.0 * area_m2 / spacing_m if spacing_m > 0.0 else 0.0
    footing_excavation_m3 = support_count * footing_length * footing_width * footing_depth
    excavation_m3 = area_m2 * excavation_depth_m + footing_excavation_m3
    formwork_m2 = support_count * 2.0 * (footing_length + footing_width) * footing_depth

    holding_down_bolts = 0
    for support in (connection_design or {}).get("base_plates", {}).get("supports", []):
        holding_down_bolts += int(
            support.get("holding_down_bolts", {}).get("layout", {}).get("bolt_count", 0) or 0
        )
    holding_down_bolts *= max(frame_count, 1)

    return {
        "schema_version": 1,
        "project": str(project.get("project_name") or "PORTAL FRAME PROJECT"),
        "inputs": {
            "surface_bed_area_m2": area_m2,
            "surface_bed_thickness_mm": thickness_mm,
            "joint_spacing_m": spacing_m,
            "excavation_below_surface_bed_m": excavation_depth_m,
            "concrete_footing_backfill_m3": footing_backfill_m3,
        },
        "items": [
            {"description": "Excavation for restricted foundations in all materials and dispose spoil material at a site provided by the Contractor - depth up to 2m", "unit": "m3", "quantity": excavation_m3, "template_row": 62},
            {"description": "Soilcrete backfill under structures as indicated on the drawings or directed by the Engineer", "unit": "m3", "quantity": footing_backfill_m3, "template_row": 66},
            {"description": "Bases - Class 30MPa/19mm concrete", "unit": "m3", "quantity": support_count * footing_volume, "template_row": 146},
            {"description": "Surface beds including thickenings", "unit": "m3", "quantity": area_m2 * thickness_mm / 1000.0, "template_row": 156},
            {"description": "Top of surface beds - power-floated finish", "unit": "m2", "quantity": area_m2, "template_row": 168},
            {"description": "Joints (D1) - Saw Cut Joint in floor", "unit": "m", "quantity": joint_length_m, "template_row": 180},
            {"description": "Ref.395 mesh reinforcement", "unit": "m2", "quantity": area_m2, "template_row": 132},
            {"description": "250 Micron damp-proof sheeting (vapour barrier) under surface beds", "unit": "m2", "quantity": area_m2, "template_row": 218},
            {"description": "Sides of bases - formwork", "unit": "m2", "quantity": formwork_m2, "template_row": 92},
            {"description": "Holding-down bolts to structural column bases", "unit": "No", "quantity": holding_down_bolts, "template_row": 210},
        ],
        "assumptions": [
            "Surface-bed area defaults to the building footprint when no separate area is entered.",
            "Saw-cut joint length is estimated as 2 x surface-bed area / joint spacing.",
            "Restricted excavation includes the entered excavation below the surface bed plus the automatically designed pad footprints to the foundation base depth.",
            "Concrete footing backfill is a direct user input and is not inferred from soil conditions.",
        ],
    }


def write_civil_boq_xlsx(
    takeoff: Mapping[str, Any],
    output_path: str | Path,
    *,
    template_path: str | Path = CIVIL_TEMPLATE,
) -> Path:
    workbook = load_workbook(template_path)
    sheet = workbook["RO Building"]
    sheet.title = "Civil and Concrete BOQ"
    item_by_row = {int(item["template_row"]): item for item in takeoff["items"]}
    for row in range(1, sheet.max_row + 1):
        unit = sheet.cell(row, 4).value
        if unit in {"m3", "m2", "m", "No", "t"}:
            for column in range(8, sheet.max_column + 1):
                sheet.cell(row, column).value = None
            sheet.cell(row, 5).value = 0.0
            sheet.cell(row, 7).value = f"=E{row}*F{row}"
    for row, item in item_by_row.items():
        sheet.cell(row, 5).value = float(item["quantity"])
        sheet.cell(row, 7).value = f"=E{row}*F{row}"
    sheet["A1"] = str(takeoff.get("project") or "PORTAL FRAME PROJECT").upper()
    sheet["A2"] = "CIVIL AND CONCRETE WORKS SCHEDULE OF QUANTITIES"
    sheet["A3"] = "RO BUILDING"
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
