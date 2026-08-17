from __future__ import annotations

import math
from pathlib import Path
import tempfile
import unittest

from reporting_workflow.boq import (
    build_structural_boq_takeoff,
    normalise_additional_items,
    write_structural_boq_xlsx,
)


def sample_snapshot() -> dict:
    return {
        "analysis": {"analysis_id": "boq-test"},
        "input_data": {
            "frame_data": [{
                "column_section_type": "H-Sections",
                "rafter_section_type": "I-Sections",
            }]
        },
        "results": {
            "project": {
                "project_name": "BOQ Test Building",
                "project_number": "PF-001",
                "designer": "Test Designer",
                "building_type": "Normal Building",
                "roof_type": "Duo Pitched",
                "building_length_mm": 48_000,
                "gable_width_mm": 16_000,
                "eaves_height_mm": 6_500,
                "apex_height_mm": 8_500,
                "rafter_spacing_mm": 6_000,
                "column_section": "203x203x46",
                "rafter_section": "457x191x74",
                "girt_section": "175x65x20x2.5",
                "girt_max_spacing_mm": 1_600,
                "wall_openings_m2": {
                    "side_1": 5,
                    "side_2": 5,
                    "gable_1": 5,
                    "gable_2": 5,
                },
            },
            "frame_summary": {
                "column_length_m": 13.0,
                "rafter_length_m": 16.4924225,
                "haunch_mass_per_frame_kg": 80.0,
                "steel_mass_breakdown": {
                    "portal_frames": {"quantity": 9},
                    "bracing": {"braced_bay_count": 2},
                    "purlins": {
                        "section": "125x50x20x2.5",
                        "total_length_m": 528.0,
                    },
                },
            },
            "bracing_design": {
                "gable_columns": [{
                    "section_type": "H-Sections",
                    "section": "203x203x46",
                    "height_mm": 8_500,
                }],
                "bracing_members": [
                    {
                        "member_type": "Roof X-brace",
                        "section_family": "Equal Angles",
                        "section": "40x40x4",
                    },
                    {
                        "member_type": "Longitudinal side-wall brace",
                        "section_family": "Equal Angles",
                        "section": "45x45x4",
                        "length_mm": 8_845,
                    },
                ],
                "roof_layout": {
                    "bay_length_mm": 6_000,
                    "roof_points": [
                        {"x_mm": 0, "y_mm": 6_500},
                        {"x_mm": 8_000, "y_mm": 8_500},
                        {"x_mm": 16_000, "y_mm": 6_500},
                    ],
                    "brace_panels": [
                        {"start_index": 0, "end_index": 1},
                        {"start_index": 1, "end_index": 2},
                    ],
                },
                "column_bracing_layout": {"members_per_wall": 2},
            },
        },
    }


def sample_connections() -> dict:
    return {
        "base_plates": {
            "supports": [
                {
                    "plate": {
                        "length_mm": 450,
                        "width_mm": 400,
                        "provided_thickness_mm": 25,
                    },
                    "holding_down_bolts": {
                        "steel_grade": "8.8",
                        "layout": {"diameter_mm": 20, "bolt_count": 4},
                    },
                },
                {
                    "plate": {
                        "length_mm": 450,
                        "width_mm": 400,
                        "provided_thickness_mm": 25,
                    },
                    "holding_down_bolts": {
                        "steel_grade": "8.8",
                        "layout": {"diameter_mm": 20, "bolt_count": 4},
                    },
                },
            ]
        },
        "haunch_connections": {
            "locations": [{
                "connection_type": "eaves_end_plate",
                "connection": {
                    "plate": {
                        "height_mm": 700,
                        "width_mm": 240,
                        "provided_thickness_mm": 20,
                    },
                    "bolts": {"diameter_mm": 24, "bolt_count": 8},
                },
            }]
        },
    }


class StructuralBoqTests(unittest.TestCase):
    def test_exact_sections_weights_cladding_and_bolts(self) -> None:
        takeoff = build_structural_boq_takeoff(
            sample_snapshot(),
            sample_connections(),
            [{
                "description": "Sundry steel allowance",
                "unit": "Sum",
                "quantity": "1",
                "rate": "12500",
            }],
        )
        items = {item["description"]: item for item in takeoff["steel_items"]}
        self.assertIn("Columns - 203x203x46", items)
        self.assertIn("Rafters - 457x191x74", items)
        self.assertAlmostEqual(
            items["Columns - 203x203x46"]["mass_kg"],
            13.0 * 9 * 46.1,
            places=6,
        )
        roof = takeoff["cladding_items"][0]
        self.assertAlmostEqual(
            roof["quantity"],
            2 * math.hypot(8, 2) * 48,
            places=6,
        )
        wall = takeoff["cladding_items"][1]
        self.assertAlmostEqual(wall["quantity"], 844.0, places=6)
        bolt_quantities = {
            item["description"]: item["quantity"]
            for item in takeoff["bolt_items"]
        }
        self.assertEqual(
            bolt_quantities[
                "Grade 8.8 M20 holding-down bolts, nuts and washers"
            ],
            72,
        )
        reference = {item["description"]: item for item in takeoff["reference_items"]}
        self.assertAlmostEqual(
            reference["Grade 8.8 precision galvanised bolts, nuts and washers"]["quantity"],
            takeoff["fabricated_steel_mass_t"] * 0.045,
            places=9,
        )
        self.assertIn("82 degrees bullnose", takeoff["cladding_items"][2]["description"])
        self.assertIn("Hot-Dipped Galvanising", reference[next(
            description for description in reference if description.startswith("Corrosion protection")
        )]["description"])
        self.assertEqual(takeoff["additional_items"][0]["rate"], 12_500)

    def test_additional_item_validation(self) -> None:
        with self.assertRaisesRegex(ValueError, "requires a description"):
            normalise_additional_items([
                {"description": "", "unit": "No", "quantity": "2"}
            ])
        with self.assertRaisesRegex(ValueError, "unit must be one of"):
            normalise_additional_items([
                {"description": "Item", "unit": "lot", "quantity": "2"}
            ])

    def test_workbook_contains_editable_rates_and_amount_formulas(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is installed with the application requirements")
        takeoff = build_structural_boq_takeoff(
            sample_snapshot(), sample_connections(), []
        )
        with tempfile.TemporaryDirectory() as directory:
            path = write_structural_boq_xlsx(
                takeoff, Path(directory) / "structural_steel_boq.xlsx"
            )
            workbook = load_workbook(path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Structural Steel BoQ"])
            sheet = workbook["Structural Steel BoQ"]
            descriptions = {
                sheet.cell(row, 3).value: row
                for row in range(1, sheet.max_row + 1)
            }
            rafter_row = descriptions["Rafters - 457x191x74"]
            self.assertEqual(sheet.cell(rafter_row, 6).value, 0)
            self.assertEqual(
                sheet.cell(rafter_row, 7).value,
                f"=E{rafter_row}*F{rafter_row}",
            )
            self.assertEqual(
                sheet.cell(sheet.max_row, 1).value,
                "TOTAL SECTION 1 CARRIED TO SUMMARY",
            )


if __name__ == "__main__":
    unittest.main()
